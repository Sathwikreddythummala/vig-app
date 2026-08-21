from fastapi import APIRouter, Request
from fastapi.responses import JSONResponse, StreamingResponse, RedirectResponse
from services.sheets_service import get_all_records, now_str, gen_id, find_row_by_id, append_row, update_row, add_audit_log
from utils.templates import templates
import io

router = APIRouter(prefix="/reports", tags=["reports"])


def get_user(request: Request):
    return request.session.get("user")


def _shift_month(ym: str, delta: int) -> str:
    """'2026-07' + delta months -> 'YYYY-MM'."""
    try:
        y, m = int(ym[:4]), int(ym[5:7])
    except (ValueError, IndexError):
        return ym
    idx = (y * 12 + (m - 1)) + delta
    return f"{idx // 12:04d}-{idx % 12 + 1:02d}"


def _num(x):
    try:
        return float(x or 0)
    except (ValueError, TypeError):
        return 0.0


def compute_vehicle_pl(expense_month: str, billing_month: str):
    """Per-vehicle P&L: revenue = billing (without GST) for billing_month,
    expenses = that vehicle's expenses for expense_month, split by type."""
    expenses = get_all_records("Expenses")
    bills = get_all_records("Billing")
    vehicles = get_all_records("Vehicles")

    def blank():
        return {"vehicle": "", "revenue": 0.0, "fuel": 0.0, "driver": 0.0, "emi": 0.0, "other": 0.0}

    # Only OWN vehicles (from the Vehicles table). Outside vehicles are excluded.
    data = {}
    for v in vehicles:
        vn = str(v.get("VehicleNumber", "")).strip()
        if vn:
            d = blank(); d["vehicle"] = vn
            data[vn] = d

    # Expenses actually dated within the month (1st -> month end). This is by ExpenseDate,
    # so driver salaries paid this month (for the previous month) are correctly included.
    for e in expenses:
        vn = str(e.get("VehicleNumber", "")).strip()
        if not vn or vn not in data:   # skip company/blank and outside vehicles
            continue
        if str(e.get("ExpenseDate", ""))[:7] != expense_month:
            continue
        d = data[vn]
        amt = _num(e.get("Amount"))
        cat = str(e.get("Category", "")).strip()
        if cat == "Fuel":
            d["fuel"] += amt
        elif cat == "Driver Expense":
            d["driver"] += amt
        elif cat == "EMI":
            d["emi"] += amt
        else:
            d["other"] += amt

    # Billing (revenue) for the billing month — WITHOUT GST (use SubTotal, pre SGST/CGST)
    for b in bills:
        vn = str(b.get("VehicleNumber", "")).strip()
        if not vn or vn not in data:   # own vehicles only; skip outside vehicles
            continue
        bm = str(b.get("PaymentMonth", "")).strip() or str(b.get("InvoiceDate", ""))[:7]
        if bm != billing_month:
            continue
        d = data[vn]
        sub = b.get("SubTotal", "")
        if str(sub).strip() == "":
            sub = _num(b.get("FixedAmount")) + _num(b.get("VariableAmount")) + _num(b.get("Tollgates")) + _num(b.get("TrafficChallan"))
        d["revenue"] += _num(sub)

    rows = []
    for vn, d in data.items():
        total_exp = d["fuel"] + d["driver"] + d["emi"] + d["other"]
        if d["revenue"] == 0 and total_exp == 0:
            continue  # no activity this period
        pl = d["revenue"] - total_exp
        rows.append({
            **d,
            "total_exp": round(total_exp, 2),
            "revenue": round(d["revenue"], 2),
            "fuel": round(d["fuel"], 2), "driver": round(d["driver"], 2),
            "emi": round(d["emi"], 2), "other": round(d["other"], 2),
            "pl": round(pl, 2),
            "margin": round((pl / d["revenue"] * 100), 1) if d["revenue"] else 0.0,
        })
    rows.sort(key=lambda x: -x["pl"])
    totals = {k: round(sum(r[k] for r in rows), 2) for k in ("revenue", "fuel", "driver", "emi", "other", "total_exp", "pl")}
    totals["margin"] = round((totals["pl"] / totals["revenue"] * 100), 1) if totals["revenue"] else 0.0
    return rows, totals


def _months(request_expense, request_billing):
    this_month = now_str()[:7]
    expense_month = request_expense or _shift_month(this_month, -1)
    billing_month = request_billing or _shift_month(expense_month, 1)
    return expense_month, billing_month


@router.get("")
async def reports_page(request: Request):
    if not get_user(request):
        return RedirectResponse("/auth/login-page")
    return templates.TemplateResponse(request=request, name="reports.html", context={"user": get_user(request)})


@router.get("/api/vehicle-pl")
async def vehicle_pl(request: Request, expense_month: str = "", billing_month: str = ""):
    if not get_user(request):
        return JSONResponse({"error": "Unauthorized"}, 401)
    em, bm = _months(expense_month, billing_month)
    rows, totals = compute_vehicle_pl(em, bm)
    return {"rows": rows, "totals": totals, "expense_month": em, "billing_month": bm}


@router.get("/api/vehicle-pl/excel")
async def vehicle_pl_excel(request: Request, expense_month: str = "", billing_month: str = ""):
    if not get_user(request):
        return JSONResponse({"error": "Unauthorized"}, 401)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    em, bm = _months(expense_month, billing_month)
    rows, totals = compute_vehicle_pl(em, bm)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vehicle P&L"
    gold = PatternFill("solid", fgColor="D4A017")
    light = PatternFill("solid", fgColor="FFF7D6")
    green = Font(color="1B7A34", bold=True)
    red = Font(color="C0392B", bold=True)
    white_bold = Font(bold=True, color="3E2723")
    hdr_font = Font(bold=True, color="3E2723", size=11)
    thin = Border(*[Side(style="thin", color="E5C76B")] * 4)
    money = '#,##0'

    ws.merge_cells("A1:I1")
    ws["A1"] = "Vigneshwara Enterprises — Vehicle Profit & Loss"
    ws["A1"].font = Font(bold=True, size=15, color="3E2723")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:I2")
    ws["A2"] = f"Expenses: {em}   |   Billing (excl. GST): {bm}"
    ws["A2"].font = Font(italic=True, size=10, color="6D4C41")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["Vehicle", "Revenue (excl GST)", "Fuel", "Driver", "EMI", "Other", "Total Expense", "Profit / Loss", "Margin %"]
    keys = ["vehicle", "revenue", "fuel", "driver", "emi", "other", "total_exp", "pl", "margin"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = hdr_font; cell.fill = gold; cell.border = thin
        cell.alignment = Alignment(horizontal="center")
    r = 5
    for row in rows:
        shade = (r % 2 == 1)
        for c, k in enumerate(keys, 1):
            cell = ws.cell(row=r, column=c, value=row[k])
            cell.border = thin
            if shade:
                cell.fill = light
            if 2 <= c <= 8:
                cell.number_format = money
            if k == "pl":
                cell.font = green if row["pl"] >= 0 else red
            if k == "margin":
                cell.number_format = '0.0"%"'
        r += 1
    # totals
    ws.cell(row=r, column=1, value="TOTAL").font = white_bold
    for c, k in enumerate(keys, 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = gold; cell.border = thin
        if c == 1:
            continue
        cell.value = totals[k]
        cell.font = white_bold
        cell.number_format = ('0.0"%"' if k == "margin" else money)
    widths = [16, 18, 12, 12, 12, 12, 15, 15, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w
    ws.freeze_panes = "A5"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=vehicle_pl_{em}.xlsx'})


def compute_mileage(month: str):
    """Per-vehicle fuel mileage. Diesel litres/amount come from Fuel entries (Diesel only)
    for the month; approved km and any litre override come from saved MileageReports."""
    fuel = get_all_records("FuelEntries")
    vehicles = get_all_records("Vehicles")
    saved = {(str(s.get("Month", "")), str(s.get("VehicleNumber", "")).strip()): s
             for s in get_all_records("MileageReports")}

    per = {}
    for v in vehicles:
        vn = str(v.get("VehicleNumber", "")).strip()
        if vn and str(v.get("VehicleStatus", "Active")).strip().lower() not in ("inactive", "sold"):
            per[vn] = {"vehicle": vn, "fuel_litres": 0.0, "diesel_amount": 0.0}

    for f in fuel:
        vn = str(f.get("VehicleNumber", "")).strip()
        if vn not in per:
            continue
        if str(f.get("FuelType", "")).strip().lower() != "diesel":
            continue
        if str(f.get("EntryDate", ""))[:7] != month:
            continue
        per[vn]["fuel_litres"] += _num(f.get("Litres"))
        per[vn]["diesel_amount"] += _num(f.get("Amount"))

    rows = []
    for vn, d in per.items():
        s = saved.get((month, vn))
        approved_km = _num(s.get("ApprovedKm")) if s else 0.0
        litres_override = (str(s.get("Litres", "")).strip() if s else "")
        litres = _num(litres_override) if litres_override != "" else round(d["fuel_litres"], 2)
        # show every own vehicle so approved km can be entered even without diesel entries
        mileage = round(approved_km / litres, 2) if litres else 0.0
        rows.append({
            "vehicle": vn,
            "fuel_litres": round(d["fuel_litres"], 2),   # actual from fuel entries
            "litres": round(litres, 2),                  # effective (override or actual)
            "diesel_amount": round(d["diesel_amount"], 2),
            "approved_km": round(approved_km, 2),
            "mileage": mileage,
        })
    rows.sort(key=lambda x: str(x["vehicle"]))
    tot_litres = sum(r["litres"] for r in rows)
    tot_km = sum(r["approved_km"] for r in rows)
    totals = {
        "fuel_litres": round(sum(r["fuel_litres"] for r in rows), 2),
        "litres": round(tot_litres, 2),
        "diesel_amount": round(sum(r["diesel_amount"] for r in rows), 2),
        "approved_km": round(tot_km, 2),
        "mileage": round(tot_km / tot_litres, 2) if tot_litres else 0.0,
    }
    return rows, totals


@router.get("/api/mileage")
async def mileage(request: Request, month: str = ""):
    if not get_user(request):
        return JSONResponse({"error": "Unauthorized"}, 401)
    month = month or now_str()[:7]
    rows, totals = compute_mileage(month)
    return {"rows": rows, "totals": totals, "month": month}


@router.post("/api/mileage/save")
async def mileage_save(request: Request):
    user = get_user(request)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, 401)
    data = await request.json()
    month = str(data.get("month", "")).strip()
    items = data.get("rows") or []
    if not month:
        return JSONResponse({"error": "Month required"}, 400)
    from services.sheets_service import build_row, SHEET_HEADERS
    existing = get_all_records("MileageReports")
    idx_map = {(str(s.get("Month", "")), str(s.get("VehicleNumber", "")).strip()): s for s in existing}
    saved = 0
    for it in items:
        vn = str(it.get("vehicle", "")).strip()
        if not vn:
            continue
        vals = {
            "Month": month, "VehicleNumber": vn,
            "ApprovedKm": str(it.get("approved_km", "") or ""),
            "Litres": str(it.get("litres", "") or ""),
            "UpdatedBy": user.get("email", ""), "UpdatedDate": now_str(),
        }
        prev = idx_map.get((month, vn))
        if prev:
            vals["MileageID"] = prev.get("MileageID", "")
            res = find_row_by_id("MileageReports", prev.get("MileageID", ""))
            if res:
                update_row("MileageReports", res[0], build_row("MileageReports", vals))
        else:
            vals["MileageID"] = gen_id("MLG")
            append_row("MileageReports", build_row("MileageReports", vals))
        saved += 1
    add_audit_log("UPDATE", "MileageReports", month, f"Mileage saved for {saved} vehicle(s) ({month})", user["email"])
    return {"success": True, "saved": saved}


@router.get("/api/mileage/excel")
async def mileage_excel(request: Request, month: str = ""):
    if not get_user(request):
        return JSONResponse({"error": "Unauthorized"}, 401)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    month = month or now_str()[:7]
    rows, totals = compute_mileage(month)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Fuel Mileage"
    gold = PatternFill("solid", fgColor="D4A017"); light = PatternFill("solid", fgColor="FFF7D6")
    hdr = Font(bold=True, color="3E2723", size=11); wb_bold = Font(bold=True, color="3E2723")
    thin = Border(*[Side(style="thin", color="E5C76B")] * 4)
    ws.merge_cells("A1:F1"); ws["A1"] = "Vigneshwara Enterprises — Fuel Mileage (Diesel)"
    ws["A1"].font = Font(bold=True, size=15, color="3E2723"); ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:F2"); ws["A2"] = f"Month: {month}"
    ws["A2"].font = Font(italic=True, size=10, color="6D4C41"); ws["A2"].alignment = Alignment(horizontal="center")
    heads = ["Vehicle", "Diesel Litres", "Diesel Amount", "Approved KM", "Mileage (km/L)"]
    keys = ["vehicle", "litres", "diesel_amount", "approved_km", "mileage"]
    for c, h in enumerate(heads, 1):
        cell = ws.cell(row=4, column=c, value=h); cell.font = hdr; cell.fill = gold; cell.border = thin
        cell.alignment = Alignment(horizontal="center")
    r = 5
    for row in rows:
        for c, k in enumerate(keys, 1):
            cell = ws.cell(row=r, column=c, value=row[k]); cell.border = thin
            if r % 2 == 1:
                cell.fill = light
            if k in ("litres", "diesel_amount", "approved_km"):
                cell.number_format = '#,##0'
            if k == "mileage":
                cell.number_format = '0.00'
        r += 1
    ws.cell(row=r, column=1, value="TOTAL / AVG").font = wb_bold
    for c, k in enumerate(keys, 1):
        cell = ws.cell(row=r, column=c); cell.fill = gold; cell.border = thin
        if c == 1:
            continue
        cell.value = totals[k]; cell.font = wb_bold
        cell.number_format = '0.00' if k == "mileage" else '#,##0'
    for i, w in enumerate([16, 14, 14, 14, 15]):
        ws.column_dimensions[chr(65 + i)].width = w
    ws.freeze_panes = "A5"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename=fuel_mileage_{month}.xlsx'})


@router.get("/api/mileage/pdf")
async def mileage_pdf(request: Request, month: str = ""):
    if not get_user(request):
        return JSONResponse({"error": "Unauthorized"}, 401)
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    month = month or now_str()[:7]
    rows, totals = compute_mileage(month)

    def rs(n):
        return f"{n:,.0f}"

    buf = io.BytesIO(); doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("t", parent=styles["Title"], textColor=colors.HexColor("#3E2723"), fontSize=16)
    sub = ParagraphStyle("s", parent=styles["Normal"], textColor=colors.HexColor("#6D4C41"), fontSize=10, alignment=1)
    elements = [Paragraph("Vigneshwara Enterprises &mdash; Fuel Mileage (Diesel)", title),
                Paragraph(f"Month: {month}", sub), Spacer(1, 10)]
    data = [["Vehicle", "Diesel Litres", "Diesel Amount", "Approved KM", "Mileage (km/L)"]]
    for row in rows:
        data.append([row["vehicle"], rs(row["litres"]), "Rs." + rs(row["diesel_amount"]), rs(row["approved_km"]), f'{row["mileage"]:.2f}'])
    data.append(["TOTAL / AVG", rs(totals["litres"]), "Rs." + rs(totals["diesel_amount"]), rs(totals["approved_km"]), f'{totals["mileage"]:.2f}'])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D4A017")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#3E2723")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5C76B")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#FFF7D6")]),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"), ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFE08A")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    elements.append(table); doc.build(elements); buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename=fuel_mileage_{month}.pdf'})


@router.get("/api/vehicle-pl/pdf")
async def vehicle_pl_pdf(request: Request, expense_month: str = "", billing_month: str = ""):
    if not get_user(request):
        return JSONResponse({"error": "Unauthorized"}, 401)
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    em, bm = _months(expense_month, billing_month)
    rows, totals = compute_vehicle_pl(em, bm)

    def rs(n):
        return f"Rs.{n:,.0f}"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=14 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("t", parent=styles["Title"], textColor=colors.HexColor("#3E2723"), fontSize=17)
    sub = ParagraphStyle("s", parent=styles["Normal"], textColor=colors.HexColor("#6D4C41"), fontSize=10, alignment=1)
    elements = [Paragraph("Vigneshwara Enterprises &mdash; Vehicle Profit &amp; Loss", title),
                Paragraph(f"Expenses: {em} &nbsp;&nbsp;|&nbsp;&nbsp; Billing (excl. GST): {bm}", sub), Spacer(1, 10)]

    header = ["Vehicle", "Revenue\n(excl GST)", "Fuel", "Driver", "EMI", "Other", "Total\nExpense", "Profit / Loss", "Margin"]
    data = [header]
    for row in rows:
        data.append([row["vehicle"], rs(row["revenue"]), rs(row["fuel"]), rs(row["driver"]),
                     rs(row["emi"]), rs(row["other"]), rs(row["total_exp"]), rs(row["pl"]), f'{row["margin"]:.1f}%'])
    data.append(["TOTAL", rs(totals["revenue"]), rs(totals["fuel"]), rs(totals["driver"]), rs(totals["emi"]),
                 rs(totals["other"]), rs(totals["total_exp"]), rs(totals["pl"]), f'{totals["margin"]:.1f}%'])

    table = Table(data, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D4A017")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#3E2723")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5C76B")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#FFF7D6")]),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFE08A")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    # colour the Profit/Loss column per row (green profit, red loss)
    for i, row in enumerate(rows, start=1):
        style.append(("TEXTCOLOR", (7, i), (7, i), colors.HexColor("#1B7A34") if row["pl"] >= 0 else colors.HexColor("#C0392B")))
    style.append(("TEXTCOLOR", (7, -1), (7, -1), colors.HexColor("#1B7A34") if totals["pl"] >= 0 else colors.HexColor("#C0392B")))
    table.setStyle(TableStyle(style))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename=vehicle_pl_{em}.pdf'})

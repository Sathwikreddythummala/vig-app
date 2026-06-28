from fastapi import APIRouter, Request, UploadFile, File, Form
from fastapi.responses import JSONResponse
from services.sheets_service import (
    get_all_records, find_row_by_id, append_row, update_row, delete_row,
    gen_id, now_str, add_audit_log,
)
from services.drive_service import upload_file
from utils.templates import templates

router = APIRouter(prefix="/drivers", tags=["drivers"])


def get_user(request: Request):
    return request.session.get("user")


def _sync_user_access(email: str, name: str, emp_type: str, access_type: str):
    if not email or not email.strip():
        return
    email = email.strip().lower()
    emp_type_norm = str(emp_type or "").strip().lower()
    if emp_type_norm == "employee" and access_type:
        users = get_all_records("Users")
        for idx, u in enumerate(users):
            if str(u.get("Email", "")).strip().lower() == email:
                from services.sheets_service import SHEET_HEADERS
                headers = SHEET_HEADERS["Users"]
                row_data = [str(u.get(h, "")) for h in headers]
                row_data[headers.index("Name")] = name
                row_data[headers.index("Role")] = access_type
                row_data[headers.index("Status")] = "Active"
                row_data[headers.index("UpdatedDate")] = now_str()
                update_row("Users", idx + 2, row_data)
                return
        append_row("Users", [gen_id("USR"), email, name, access_type, "Active", now_str(), now_str()])
    elif emp_type_norm == "driver":
        users = get_all_records("Users")
        for idx, u in enumerate(users):
            if str(u.get("Email", "")).strip().lower() == email:
                from services.sheets_service import SHEET_HEADERS
                headers = SHEET_HEADERS["Users"]
                row_data = [str(u.get(h, "")) for h in headers]
                row_data[headers.index("Status")] = "Inactive"
                row_data[headers.index("UpdatedDate")] = now_str()
                update_row("Users", idx + 2, row_data)
                return


@router.get("")
async def drivers_page(request: Request):
    user = get_user(request)
    if not user:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/auth/login-page")
    return templates.TemplateResponse(request=request, name="drivers.html", context={"user": user})


@router.get("/api/list")
async def list_drivers(request: Request):
    user = get_user(request)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, 401)
    drivers = get_all_records("Drivers")
    drivers.sort(key=lambda d: str(d.get("AssignedVehicle", "") or "ZZZ"))
    return {"drivers": drivers}


@router.get("/api/salaries")
async def salaries_api(request: Request, month: str = ""):
    user = get_user(request)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, 401)
    drivers = get_all_records("Drivers")
    expenses = get_all_records("Expenses")
    result = []
    for d in drivers:
        if str(d.get("Status", "Active")).strip().lower() == "inactive":
            continue
        name = str(d.get("DriverName", "")).strip()
        profile_salary = float(d.get("Salary", 0) or 0)
        drv_expenses = [e for e in expenses if str(e.get("DriverName", "")).strip() == name]
        if month:
            drv_expenses = [e for e in drv_expenses if e.get("ForMonth", "") == month or str(e.get("ExpenseDate", ""))[:7] == month]
        salary_paid = sum(float(e.get("Amount", 0) or 0) for e in drv_expenses if e.get("SubCategory") == "Salary")
        advance = sum(float(e.get("Amount", 0) or 0) for e in drv_expenses if e.get("SubCategory") == "Advance")
        meals = sum(float(e.get("Amount", 0) or 0) for e in drv_expenses if e.get("SubCategory") == "Meals")
        other = sum(float(e.get("Amount", 0) or 0) for e in drv_expenses if e.get("SubCategory") not in ("Salary", "Advance", "Meals"))
        net_payable = profile_salary - advance - meals - other
        result.append({
            "DriverID": d.get("DriverID", ""),
            "DriverName": name,
            "EmployeeType": d.get("EmployeeType", "Driver"),
            "AssignedVehicle": d.get("AssignedVehicle", ""),
            "ProfileSalary": profile_salary,
            "SalaryPaid": salary_paid,
            "Advance": advance,
            "Meals": meals,
            "Other": other,
            "NetPayable": net_payable,
        })
    result.sort(key=lambda x: x["DriverName"])
    totals = {
        "ProfileSalary": sum(r["ProfileSalary"] for r in result),
        "SalaryPaid": sum(r["SalaryPaid"] for r in result),
        "Advance": sum(r["Advance"] for r in result),
        "Meals": sum(r["Meals"] for r in result),
        "Other": sum(r["Other"] for r in result),
        "NetPayable": sum(r["NetPayable"] for r in result),
    }
    return {"salaries": result, "totals": totals, "month": month}


@router.get("/api/salaries/export")
async def export_salaries_excel(request: Request, month: str = ""):
    user = get_user(request)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, 401)
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return JSONResponse({"error": "openpyxl not installed"}, 500)
    data = await salaries_api(request, month)
    salaries = data.body
    import json
    body = json.loads(salaries)
    rows = body["salaries"]
    totals = body["totals"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Salaries {month or 'All'}"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="D4A017", end_color="D4A017", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    total_font = Font(bold=True, size=11)
    currency_fmt = '#,##0'
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Salary Report — {month or 'All Months'}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    headers = ["#", "Name", "Type", "Vehicle", "Salary", "Advance", "Meals", "Other", "Net Payable"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for i, r in enumerate(rows):
        row_num = i + 4
        ws.cell(row=row_num, column=1, value=i+1).border = thin_border
        ws.cell(row=row_num, column=2, value=r["DriverName"]).border = thin_border
        ws.cell(row=row_num, column=3, value=r["EmployeeType"]).border = thin_border
        ws.cell(row=row_num, column=4, value=r["AssignedVehicle"]).border = thin_border
        for col, key in [(5, "ProfileSalary"), (6, "Advance"), (7, "Meals"), (8, "Other"), (9, "NetPayable")]:
            cell = ws.cell(row=row_num, column=col, value=r[key])
            cell.number_format = currency_fmt
            cell.border = thin_border
    total_row = len(rows) + 4
    ws.cell(row=total_row, column=1, value="").border = thin_border
    ws.cell(row=total_row, column=2, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=2).border = thin_border
    ws.cell(row=total_row, column=3, value="").border = thin_border
    ws.cell(row=total_row, column=4, value="").border = thin_border
    for col, key in [(5, "ProfileSalary"), (6, "Advance"), (7, "Meals"), (8, "Other"), (9, "NetPayable")]:
        cell = ws.cell(row=total_row, column=col, value=totals[key])
        cell.number_format = currency_fmt
        cell.font = total_font
        cell.border = thin_border
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    for c in ["E", "F", "G", "H", "I"]:
        ws.column_dimensions[c].width = 14
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    filename = f"Salaries_{month or 'All'}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/api/{driver_id}")
async def get_driver(request: Request, driver_id: str):
    user = get_user(request)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, 401)
    result = find_row_by_id("Drivers", driver_id)
    if not result:
        return JSONResponse({"error": "Driver not found"}, 404)
    _, record = result
    docs = get_all_records("Documents")
    driver_docs = [d for d in docs if d.get("EntityType") == "Driver" and d.get("EntityID") == driver_id]
    expenses = get_all_records("Expenses")
    driver_expenses = [e for e in expenses if str(e.get("DriverName", "")) == str(record.get("DriverName", ""))]
    return {"driver": record, "documents": driver_docs, "expenses": driver_expenses}


@router.post("/api/add")
async def add_driver(request: Request):
    user = get_user(request)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, 401)
    data = await request.json()
    drivers = get_all_records("Drivers")
    lic = str(data.get("DrivingLicenseNumber", "")).strip().upper()
    if lic:
        for d in drivers:
            if str(d.get("DrivingLicenseNumber", "")).strip().upper() == lic:
                return JSONResponse({"error": "Driving license number already exists"}, 400)
    did = gen_id("DRV")
    row = [
        did,
        data.get("EmployeeType", "Driver"),
        data.get("DriverName", ""),
        data.get("Email", ""),
        data.get("MobileNumber", ""),
        data.get("EmergencyContact", ""),
        data.get("Address", ""),
        data.get("AadhaarNumber", ""),
        data.get("DrivingLicenseNumber", ""),
        data.get("LicenseExpiryDate", ""),
        data.get("BankName", ""),
        data.get("AccountNumber", ""),
        data.get("IFSCCode", ""),
        data.get("Salary", ""),
        data.get("JoiningDate", ""),
        data.get("Status", "Active"),
        data.get("ExitDate", ""),
        data.get("AssignedVehicle", ""),
        now_str(),
        now_str(),
    ]
    append_row("Drivers", row)
    new_vehicle = str(data.get("AssignedVehicle", "")).strip()
    if new_vehicle:
        from services.sheets_service import SHEET_HEADERS
        vehicles = get_all_records("Vehicles")
        veh_headers = SHEET_HEADERS["Vehicles"]
        drv_idx = veh_headers.index("DefaultDriver")
        veh_updated_idx = veh_headers.index("UpdatedDate")
        for idx, v in enumerate(vehicles):
            if str(v.get("VehicleNumber", "")).strip() == new_vehicle:
                veh_row = [v.get(h, "") for h in veh_headers]
                veh_row[drv_idx] = data.get("DriverName", "")
                veh_row[veh_updated_idx] = now_str()
                update_row("Vehicles", idx + 2, veh_row)
                break
    _sync_user_access(data.get("Email", ""), data.get("DriverName", ""), data.get("EmployeeType", "Driver"), data.get("AccessType", "viewer"))
    add_audit_log("CREATE", "Drivers", did, f"{data.get('EmployeeType','Driver')} {data.get('DriverName','')} added", user["email"])
    return {"success": True, "driver_id": did}


@router.put("/api/{driver_id}")
async def update_driver(request: Request, driver_id: str):
    user = get_user(request)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, 401)
    data = await request.json()
    result = find_row_by_id("Drivers", driver_id)
    if not result:
        return JSONResponse({"error": "Driver not found"}, 404)
    row_num, existing = result
    lic = str(data.get("DrivingLicenseNumber", "")).strip().upper()
    if lic:
        drivers = get_all_records("Drivers")
        for d in drivers:
            if str(d.get("DrivingLicenseNumber", "")).strip().upper() == lic and str(d.get("DriverID", "")) != driver_id:
                return JSONResponse({"error": "Driving license number already exists"}, 400)
    row = [
        driver_id,
        data.get("EmployeeType", existing.get("EmployeeType", "Driver")),
        data.get("DriverName", ""),
        data.get("Email", existing.get("Email", "")),
        data.get("MobileNumber", ""),
        data.get("EmergencyContact", ""),
        data.get("Address", ""),
        data.get("AadhaarNumber", ""),
        data.get("DrivingLicenseNumber", ""),
        data.get("LicenseExpiryDate", ""),
        data.get("BankName", ""),
        data.get("AccountNumber", ""),
        data.get("IFSCCode", ""),
        data.get("Salary", ""),
        data.get("JoiningDate", ""),
        data.get("Status", data.get("Status", "Active")),
        data.get("ExitDate", existing.get("ExitDate", "")),
        data.get("AssignedVehicle", ""),
        existing.get("CreatedDate", now_str()),
        now_str(),
    ]
    exit_date = str(data.get("ExitDate", "")).strip()
    if exit_date:
        row[15] = "Inactive"
    update_row("Drivers", row_num, row)
    old_vehicle = str(existing.get("AssignedVehicle", "")).strip()
    new_vehicle = str(data.get("AssignedVehicle", "")).strip()
    driver_name = str(data.get("DriverName", "")).strip()
    if old_vehicle != new_vehicle:
        from services.sheets_service import SHEET_HEADERS
        vehicles = get_all_records("Vehicles")
        veh_headers = SHEET_HEADERS["Vehicles"]
        drv_idx = veh_headers.index("DefaultDriver")
        veh_updated_idx = veh_headers.index("UpdatedDate")
        if old_vehicle:
            for idx, v in enumerate(vehicles):
                if str(v.get("VehicleNumber", "")).strip() == old_vehicle and str(v.get("DefaultDriver", "")).strip() == driver_name:
                    veh_row = [v.get(h, "") for h in veh_headers]
                    veh_row[drv_idx] = ""
                    veh_row[veh_updated_idx] = now_str()
                    update_row("Vehicles", idx + 2, veh_row)
                    break
        if new_vehicle:
            for idx, v in enumerate(vehicles):
                if str(v.get("VehicleNumber", "")).strip() == new_vehicle:
                    veh_row = [v.get(h, "") for h in veh_headers]
                    veh_row[drv_idx] = driver_name
                    veh_row[veh_updated_idx] = now_str()
                    update_row("Vehicles", idx + 2, veh_row)
                    break
    _sync_user_access(data.get("Email", existing.get("Email", "")), data.get("DriverName", ""), data.get("EmployeeType", existing.get("EmployeeType", "Driver")), data.get("AccessType", "viewer"))
    add_audit_log("UPDATE", "Drivers", driver_id, f"{data.get('EmployeeType','Driver')} {data.get('DriverName','')} updated", user["email"])
    return {"success": True}


@router.delete("/api/{driver_id}")
async def delete_driver_api(request: Request, driver_id: str):
    user = get_user(request)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, 401)
    result = find_row_by_id("Drivers", driver_id)
    if not result:
        return JSONResponse({"error": "Driver not found"}, 404)
    row_num, record = result
    delete_row("Drivers", row_num)
    add_audit_log("DELETE", "Drivers", driver_id, f"Driver {record.get('DriverName','')} deleted", user["email"])
    return {"success": True}


@router.post("/api/{driver_id}/upload")
async def upload_driver_doc(
    request: Request,
    driver_id: str,
    doc_type: str = Form(...),
    file: UploadFile = File(...),
):
    user = get_user(request)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, 401)
    content = await file.read()
    result = upload_file(content, file.filename, file.content_type or "application/octet-stream", "Drivers", doc_type)
    doc_id = gen_id("DOC")
    append_row("Documents", [doc_id, "Driver", driver_id, doc_type, file.filename, result["view_url"], now_str()])
    add_audit_log("UPLOAD", "Documents", doc_id, f"Uploaded {doc_type} for driver {driver_id}", user["email"])
    return {"success": True, "document": result, "doc_id": doc_id}


@router.get("/salaries")
async def salaries_page(request: Request):
    user = get_user(request)
    if not user:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/auth/login-page")
    return templates.TemplateResponse(request=request, name="salaries.html", context={"user": user})


@router.get("/details/{driver_id}")
async def driver_details_page(request: Request, driver_id: str):
    user = get_user(request)
    if not user:
        from fastapi.responses import RedirectResponse
        return RedirectResponse("/auth/login-page")
    return templates.TemplateResponse(request=request, name="driver_details.html", context={"user": user, "driver_id": driver_id})
